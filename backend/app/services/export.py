from __future__ import annotations

import io
from typing import Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from openpyxl import Workbook
from openpyxl.styles import Font


def generate_pdf(final_output: dict[str, Any], constraints: dict[str, Any]) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.6 * inch, bottomMargin=0.6 * inch)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "SpotOnTitle", parent=styles["Heading1"], fontSize=22, spaceAfter=6, textColor=colors.HexColor("#1d1d1f")
    )
    subtitle_style = ParagraphStyle(
        "SpotOnSubtitle", parent=styles["Normal"], fontSize=11, textColor=colors.HexColor("#86868b"), spaceAfter=20
    )
    section_style = ParagraphStyle(
        "SpotOnSection", parent=styles["Heading2"], fontSize=16, spaceBefore=24, spaceAfter=10,
        textColor=colors.HexColor("#FF4F00"),
    )
    item_name_style = ParagraphStyle(
        "ItemName", parent=styles["Heading3"], fontSize=12, spaceBefore=10, spaceAfter=4,
        textColor=colors.HexColor("#1d1d1f"),
    )
    body_style = ParagraphStyle(
        "ItemBody", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#424245"), spaceAfter=2
    )

    elements: list = []

    # Header
    origin = constraints.get("origin", "")
    destination = constraints.get("destination", "")
    departing = constraints.get("departing_date", "")
    returning = constraints.get("returning_date", "")
    trip_label = f"{origin} to {destination}" if origin and destination else "Your Trip"
    date_label = departing
    if returning:
        date_label += f" - {returning}"

    elements.append(Paragraph("Spot On", title_style))
    elements.append(Paragraph(f"{trip_label} | {date_label}" if date_label else trip_label, subtitle_style))

    # Sections
    section_configs = [
        ("Flights", final_output.get("flights", []), ["airline", "route", "trip_type", "price_range", "url"]),
        ("Car Rentals", final_output.get("car_rentals", []), ["provider", "vehicle_class", "price_per_day", "pickup_location", "url"]),
        ("Hotels", final_output.get("hotels", []), ["name", "price_per_night", "area", "why_recommended", "url"]),
        ("Dining", final_output.get("restaurants", []), ["name", "cuisine", "price_range", "area", "why_recommended", "url"]),
        ("Must-See Spots", final_output.get("travel_spots", []), ["name", "kind", "area", "why_recommended", "url"]),
    ]

    for section_title, items, fields in section_configs:
        if not items:
            continue
        elements.append(Paragraph(section_title, section_style))
        for item in items:
            name = item.get("name") or item.get("provider") or item.get("airline") or "—"
            elements.append(Paragraph(name, item_name_style))
            for field in fields:
                if field in ("name", "provider", "airline"):
                    continue
                val = item.get(field)
                if val:
                    label = field.replace("_", " ").title()
                    elements.append(Paragraph(f"<b>{label}:</b> {val}", body_style))
        elements.append(Spacer(1, 12))

    # References section
    refs = final_output.get("references", [])
    if refs:
        elements.append(Paragraph("References", section_style))
        for ref in refs:
            title = ref.get("title") or ref.get("url", "—")
            url = ref.get("url", "")
            section = ref.get("section", "")
            label = f" [{section}]" if section else ""
            if url:
                elements.append(
                    Paragraph(f'<a href="{url}" color="#0066cc">{title}</a>{label}', body_style)
                )
            else:
                elements.append(Paragraph(f"{title}{label}", body_style))
        elements.append(Spacer(1, 12))

    doc.build(elements)
    return buf.getvalue()


def generate_xlsx(final_output: dict[str, Any], constraints: dict[str, Any]) -> bytes:
    wb = Workbook()
    bold = Font(bold=True)

    sheet_configs = [
        ("Restaurants", final_output.get("restaurants", []),
         ["name", "cuisine", "price_range", "rating", "area", "operating_hours", "why_recommended", "menu_url", "reservation_url", "url"]),
        ("Attractions", final_output.get("travel_spots", []),
         ["name", "kind", "area", "operating_hours", "admission_price", "estimated_duration_min", "why_recommended", "reservation_url", "url"]),
        ("Hotels", final_output.get("hotels", []),
         ["name", "price_per_night", "area", "amenities", "why_recommended", "url"]),
        ("Car Rentals", final_output.get("car_rentals", []),
         ["provider", "vehicle_class", "price_per_day", "pickup_location", "operating_hours", "url"]),
        ("Flights", final_output.get("flights", []),
         ["airline", "route", "trip_type", "price_range", "url"]),
    ]

    first = True
    for sheet_name, items, columns in sheet_configs:
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        # Header row
        headers = [col.replace("_", " ").title() for col in columns]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = bold

        # Data rows
        for item in items:
            row = []
            for col in columns:
                val = item.get(col, "")
                if isinstance(val, list):
                    val = ", ".join(str(v) for v in val)
                row.append(val if val is not None else "")
            ws.append(row)

        # Auto-width columns (rough estimate)
        for col_idx, col_name in enumerate(columns, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(15, len(col_name) + 5)

    # References sheet
    refs = final_output.get("references", [])
    if refs:
        ws = wb.create_sheet(title="References")
        ref_columns = ["title", "section", "url"]
        headers = [col.replace("_", " ").title() for col in ref_columns]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = bold
        for ref in refs:
            ws.append([ref.get("title", ""), ref.get("section", ""), ref.get("url", "")])
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
